##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import openpyxl
from django.contrib import messages
from django.contrib.auth.decorators import login_required, permission_required
from django.core.urlresolvers import reverse
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404
from django.utils.translation import ugettext_lazy as _
from django.views.decorators.http import require_http_methods

from internship import models as mdl
from internship.models.cohort import Cohort
from internship.models.internship_offer import InternshipOffer
from internship.models.internship_speciality import InternshipSpeciality
from internship.models.organization import Organization
from internship.models.period import Period
from internship.models.period_internship_places import PeriodInternshipPlaces


@require_http_methods(['POST'])
@login_required
@permission_required('internship.is_internship_manager', raise_exception=True)
def upload_places_file(request, cohort_id):
    cohort = get_object_or_404(Cohort, pk=cohort_id)
    file_name = request.FILES['file']
    if file_name is not None:
        if ".xls" not in str(file_name):
            messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
        else:
            _save_xls_place(file_name, request.user, cohort)

    return HttpResponseRedirect(reverse('internships_places', kwargs={
        'cohort_id': cohort.id
    }))


def _save_xls_place(file_name, user, cohort):
    workbook = openpyxl.load_workbook(file_name, read_only=True)
    worksheet = workbook.active
    col_reference = 0
    col_name = 1
    col_address = 2
    col_postal_code = 3
    col_city = 4
    col_country = 5
    col_url = 6

    # Iterates over the lines of the spreadsheet.
    for count, row in enumerate(worksheet.rows):
        if row[col_reference].value is None \
                or row[col_reference].value == 0 \
                or not _is_registration_id(row[col_reference].value):
            continue

        reference = ""
        if row[col_reference].value < 10:
            reference = "0"+str(row[col_reference].value)
        else:
            reference = str(row[col_reference].value)

        place = mdl.organization.search(reference=reference)
        if place:
            organization = mdl.organization.find_by_id(place[0].id)
        else:
            organization = mdl.organization.Organization(cohort=cohort)

        if row[col_reference].value:
            reference = ""
            if int(row[col_reference].value) < 10:
                reference = "0"+str(row[col_reference].value)
            else:
                reference = str(row[col_reference].value)
            organization.reference = reference
        else:
            organization.reference = None

        if row[col_name].value:
            organization.name = row[col_name].value
            organization.acronym = row[col_name].value[:14]
        else:
            organization.name = None
            organization.acronym = None

        if row[col_url].value:
            organization.website = row[col_url].value
        else:
            organization.website = ""

        organization.type = "service partner"

        organization.save()

        if place:
            organization_address = mdl.organization_address.search(organization=organization)
            if not organization_address:
                organization_address = mdl.organization_address.OrganizationAddress()
            else:
                organization_address = organization_address[0]
        else:
            organization_address = mdl.organization_address.OrganizationAddress()

        if organization:
            organization_address.label = "Addr"+organization.name[:14]
        else:
            organization_address.label = " "

        if row[col_address].value:
            organization_address.location = row[col_address].value
        else:
            organization_address.location = " "

        if row[col_postal_code].value:
            organization_address.postal_code = row[col_postal_code].value
        else:
            organization_address.postal_code = " "

        if row[col_city].value:
            organization_address.city = row[col_city].value
        else:
            organization_address.city = " "

        if row[col_country].value:
            organization_address.country = row[col_country].value
        else:
            organization_address.country = " "
        organization_address.organization = organization
        organization_address.latitude = None
        organization_address.longitude = None
        organization_address.save()


def _is_registration_id(registration_id):
    try:
        int(registration_id)
        return True
    except ValueError:
        return False


@require_http_methods(['POST'])
@login_required
@permission_required('internship.is_internship_manager', raise_exception=True)
def upload_internships_file(request, cohort_id):
    cohort = get_object_or_404(Cohort, pk=cohort_id)
    file_name = request.FILES['file']
    if file_name is not None:
        if ".xls" not in str(file_name):
            messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
        else:
            save_xls_internships(file_name, cohort)
            # __save_xls_internships(request, file_name, request.user)

    return HttpResponseRedirect(reverse('internships', kwargs={
        'cohort_id': cohort.id,
    }))


def save_xls_internships(filename, cohort):
    workbook = openpyxl.load_workbook(filename, read_only=True)
    worksheet = workbook.active

    col_reference, col_spec, col_master = range(0, 3)

    def validated_rows(iterable):
        for row in iterable:
            reference = row[col_reference].value
            if reference and _is_registration_id(reference):
                if row[col_spec].value is not None:
                    yield row

    for row in validated_rows(worksheet.rows):
        reference = '%02d' % row[col_reference].value
        organization = Organization.objects.filter(cohort=cohort, reference=reference).first()
        if organization is None:
            continue

        spec_value = row[col_spec].value.replace(' ', '').replace('*', '')
        master_value = row[col_master].value

        specialities = InternshipSpeciality.objects.filter(cohort=cohort, acronym__icontains=spec_value)

        number_places = sum(int(row[x].value) for x in range(3, 7) if row[x].value)

        for speciality in specialities:
            parameters = {
                'cohort': cohort,
                'speciality': speciality,
                'organization': organization,
            }
            tmp_offer = InternshipOffer.objects.filter(**parameters).first()

            offer = tmp_offer if tmp_offer else InternshipOffer(**parameters)

            offer.title = speciality.name
            offer.maximum_enrollments = number_places
            offer.master = master_value
            offer.selectable = True
            offer.save()

            number_of_periods = 9
            for x in range(3, 7):
                period_name = 'P%d' % (number_of_periods)
                number_of_periods += 1

                period = Period.objects.filter(cohort=cohort, name=period_name).first()

                relation = PeriodInternshipPlaces.objects.filter(period=period, internship_offer=offer).first()

                if not relation:
                    relation = PeriodInternshipPlaces()

                relation.period = period
                relation.internship_offer = offer

                value = row[x].value
                relation.number_places = 0 if value is None else int(value)

                relation.save()


@require_http_methods(['POST'])
@login_required
@permission_required('internship.is_internship_manager', raise_exception=True)
def upload_masters_file(request, cohort_id):
    cohort = get_object_or_404(Cohort, pk=cohort_id)
    file_name = request.FILES['file']
    if file_name is not None:
        if ".xls" not in str(file_name):
            messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
        else:
            __save_xls_masters(request, file_name, request.user)

    return HttpResponseRedirect(reverse('internships_masters', kwargs={
        'cohort_id': cohort.id
    }))


@login_required
def __save_xls_masters(request, file_name, user):
    workbook = openpyxl.load_workbook(file_name, read_only=True)
    worksheet = workbook.active

    col_reference = 2
    col_firstname = 3
    col_lastname = 4
    col_mail = 7
    col_organization_reference = 6
    col_civility = 0
    col_mastery = 1
    col_speciality = 5

    # Iterates over the lines of the spreadsheet.
    for count, row in enumerate(worksheet.rows):
        check_reference = str(row[col_reference].value).strip(' ')
        if check_reference == "":
            check_reference = "000"
        if check_reference is None:
            check_reference = "000"

        if not _is_registration_id(check_reference):
            continue

        if row[col_firstname].value and row[col_lastname].value:
            master_check = mdl.internship_master.search(first_name=row[col_firstname].value,
                                                        last_name=row[col_lastname].value)
            if len(master_check) == 0:
                master = mdl.internship_master.InternshipMaster()
            else:
                master = master_check[0]

            if row[col_organization_reference].value:
                reference = ""
                check_reference = row[col_organization_reference].value.strip(' ')
                if check_reference != "":
                    if check_reference[0][0] != "0":
                        if int(check_reference) < 10:
                            reference = "0"+str(check_reference)
                        else:
                            reference = str(check_reference)
                    else:
                        reference = str(check_reference)

                    organization = mdl.organization.search(reference=reference)
                    master.organization = organization[0]
                else:
                    master.organization = None
            if row[col_firstname].value:
                master.first_name = row[col_firstname].value
            else:
                master.first_name = " "

            if row[col_lastname].value:
                master.last_name = row[col_lastname].value
            else:
                master.last_name = " "

            if row[col_reference].value:
                master.reference = row[col_reference].value
            else:
                master.reference = " "

            if row[col_civility].value:
                master.civility = row[col_civility].value
            else:
                master.civility = " "

            if row[col_mastery].value:
                master.type_mastery = row[col_mastery].value
            else:
                master.type_mastery = " "

            if row[col_speciality].value:
                master.speciality = row[col_speciality].value
            else:
                master.speciality = " "

            master.save()
